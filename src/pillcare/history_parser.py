"""Parse 심평원 '내가 먹는 약' encrypted XLS into MedRecord list."""

import io
from dataclasses import dataclass
from pathlib import Path

import msoffcrypto
import openpyxl


@dataclass
class MedRecord:
    seq: int
    drug_name: str
    drug_class: str
    ingredient: str
    drug_code: str
    unit: str
    dose_per_time: float
    times_per_day: int
    duration_days: int
    safety_letter: str
    antithrombotic: str
    department: str


def parse_history_xls(path: Path, password: str, department: str) -> list[MedRecord]:
    """Decrypt and parse a 심평원 개인투약이력 XLS file."""
    with open(path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        decrypted = io.BytesIO()
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
        decrypted.seek(0)

    wb = openpyxl.load_workbook(decrypted, read_only=True)
    ws = wb.active

    header_row_idx = None
    headers: list[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        if "번호" in cells and "제품명" in cells:
            header_row_idx = idx
            headers = cells
            break

    if header_row_idx is None:
        wb.close()
        return []

    col_map = {name: i for i, name in enumerate(headers)}
    records: list[MedRecord] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cells = list(row)
        seq_val = cells[col_map["번호"]]
        if seq_val is None:
            continue
        try:
            seq = int(seq_val)
        except ValueError, TypeError:
            continue

        def g(col_name: str) -> str:
            idx = col_map.get(col_name)
            if idx is None or idx >= len(cells):
                return ""
            return str(cells[idx]).strip() if cells[idx] is not None else ""

        def gf(col_name: str) -> float:
            v = g(col_name)
            try:
                return float(v)
            except ValueError, TypeError:
                return 0.0

        def gi(col_name: str) -> int:
            v = g(col_name)
            try:
                return int(float(v))
            except ValueError, TypeError:
                return 0

        records.append(
            MedRecord(
                seq=seq,
                drug_name=g("제품명"),
                drug_class=g("약효분류"),
                ingredient=g("성분명"),
                drug_code=g("약품코드"),
                unit=g("단위"),
                dose_per_time=gf("1회 투약량"),
                times_per_day=gi("1일 투여횟수"),
                duration_days=gi("총 투약일수"),
                safety_letter=g("안전성 서한(속보)"),
                antithrombotic=g("항혈전제 여부"),
                department=department,
            )
        )

    wb.close()
    return records
